import boto3
import pandas as pd
import re
from unidecode import unidecode
import pandas as pd
import re
from unidecode import unidecode
from openpyxl import load_workbook
import os
import warnings
import datetime
def AWS(imagen):
    fecha_actual=fecha_actual = datetime.datetime.now().strftime("%d_%m_%Y")
    warnings.simplefilter(action='ignore', category=FutureWarning)
    ruta_archivo = "output\salida"
    nombre_archivo = os.path.join(ruta_archivo, "1. DECLARACION DE PRODUCCION_"+fecha_actual+".xlsx")
    pd.set_option('mode.chained_assignment', None)
    # Nombre del bucket de Amazon S3
    bucket_name = 'abretesesamo'
    aws_access_key_id = 'AKIAR5DK3U4IKLOASTL4'
    aws_secret_access_key = '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'
    region_name = 'us-east-1'  # Ajusta la región según tu configuración
    # Crear cliente de Amazon Textract

    # Nombre del archivo en el bucket
    #document_name = 'DECLARACION PRODUCCION - FACTURA TUNXHO SAS - 05 DE MAYO DE 2023.pdf'

    arc= "s"
    # Nombre del archivo local que deseas cargar
    ruta_archivo = r"input\\zborrar\\borrar1\\"


    documentName = ruta_archivo + imagen

    # Nombre del archivo en el bucket de S3
    s3_file_name = 'decpro.jpg'

    # Nombre del bucket de S3
    bucket_name = 'abretesesamo'

    # Crear una instancia del cliente de S3
    s3_client = boto3.client('s3',
                            aws_access_key_id=aws_access_key_id,
                            aws_secret_access_key=aws_secret_access_key,
                            region_name=region_name)

    # Cargar el archivo en el bucket
    s3_client.upload_file(documentName, bucket_name, s3_file_name)

    print("Archivo cargado exitosamente en el bucket de S3.")

    # Credenciales de AWS

    textract = boto3.client('textract', region_name=region_name, aws_access_key_id=aws_access_key_id, aws_secret_access_key=aws_secret_access_key)

    # Iniciar el análisis del documento con formularios
    response = textract.start_document_analysis(
        DocumentLocation={
            'S3Object': {
                'Bucket': bucket_name,
                'Name': s3_file_name
            }
        },
        FeatureTypes=['FORMS']
    )

    JobId=response['JobId']

    response = textract.get_document_analysis(JobId=JobId)
    while response['JobStatus'] != 'SUCCEEDED':
        response = textract.get_document_analysis(JobId=JobId)
        

    eliminar = s3_client.delete_object(Bucket=bucket_name, Key=s3_file_name)

    fields = []
    rows = []
    text= []
    word=[]
    for block in response['Blocks']:
        if block['BlockType'] == 'WORD':
            word.append((block['Id'],block['Text']))
        dictionary = {word[i][0]: word[i][1] for i in range(0, len(word), 1)}
        if block['BlockType'] == 'KEY_VALUE_SET':
            for entity in block['EntityTypes']:
                if entity == 'KEY':
                    try:
                        nuevo_vector = [dictionary.get(key, key) for key in block['Relationships'][1]['Ids']]
                        fields.append((nuevo_vector,block['Relationships'][0]['Ids'],block['Page']))
                    except:
                        pass
                elif entity == 'VALUE':
                    try:
                        nuevo_vector = [dictionary.get(key, key) for key in block['Relationships'][0]['Ids']]
                        rows.append((block['Id'],nuevo_vector,block['Page']))
                    except:
                        pass
    df1 = pd.DataFrame(fields, columns=['key','Id_Value','PageA'])
    df2 = pd.DataFrame(rows, columns=['Id_Value','key','PageB'])

    for i in range(len(df1)):
        df1['Id_Value'][i] = df1['Id_Value'][i][0]

    merge_df = pd.merge(df1, df2, on='Id_Value', how='left')
    merge_df = merge_df[['key_x','key_y','PageA']]
    merged_df = merge_df.astype(str)

    def verificar_coincidencia(cadena):
        i = 1
        cadena_sin_tildes = unidecode(cadena)
        coincidencias = sum(1 for palabra in palabras_etiqueta if re.search(r'\b{}\b'.format(re.escape(palabra.strip())), cadena_sin_tildes, re.IGNORECASE))
        if len(palabras_etiqueta) > 1:
            i = 2
        return coincidencias >= i


    df_decprod_dict = {}
    etiqueta_names = ['NOMBRES', 'APELLIDOS', 'C.C', 'cantidad, mineral','UNIDAD, MEDIDA','MUNICIPIO', 'FECHA, VENTA', 'actividad, economica']
    columnas_decpro = ['1. Nombre', '2. Apellido', '3. Cc', '4. Cantidad', '5. Unidad', '6. Municipio', '7. Fecha', '8. Act']

    max_length = 0  # Longitud máxima de los arrays

    for etiqueta_name, decpro in zip(etiqueta_names, columnas_decpro):
        palabras_etiqueta = etiqueta_name.split(',')
        filas_etiqueta = merged_df.loc[((merged_df['key_x'].apply(verificar_coincidencia)) | (merged_df['key_y'].apply(verificar_coincidencia)))]
        filas_etiqueta.loc[:, 'key_y'] = filas_etiqueta['key_y'].str.replace("\\['", '', regex=True).str.replace("\\']", '', regex=True).str.replace("\\'", '', regex=True)
        if etiqueta_name != 'cantidad, mineral':
            filas_etiqueta.loc[:, 'key_y'] = filas_etiqueta['key_y'].str.replace(',', ' ', regex=True)
        valores_etiqueta = filas_etiqueta['key_y'].tolist()
        df_decprod_dict[decpro] = valores_etiqueta
        max_length = max(max_length, len(valores_etiqueta))  # Actualizar la longitud máxima

    # Rellenar los arrays con None si tienen una longitud menor que la máxima
    for decpro, valores_etiqueta in df_decprod_dict.items():
        if len(valores_etiqueta) < max_length:
            valores_etiqueta += ['x'] * (max_length - len(valores_etiqueta))

    # Crear el DataFrame a partir del diccionario
    df_decprod = pd.DataFrame(df_decprod_dict)


    # Cargar el DataFrame existente desde el archivo Excel si ya existe
    try:
        workbook = load_workbook(nombre_archivo)
        writer = pd.ExcelWriter(nombre_archivo, engine='openpyxl')
        writer.book = workbook
    except FileNotFoundError:
        writer = pd.ExcelWriter(nombre_archivo, engine='xlsxwriter')

    # Leer el número de filas existentes en la hoja de cálculo
    sheet_name = writer.book.active.title if writer.book.sheetnames else 'Sheet1'
    startrow = writer.book[sheet_name].max_row if sheet_name in writer.book.sheetnames else 0

    # Verificar si la primera fila está vacía
    is_empty = startrow == 0

    # Renombrar las columnas del DataFrame
    df_decprod.columns = ['1. Nombre', '2. Apellido', '3. Cc', '4. Cantidad', '5. Unidad', '6. Municipio', '7. Fecha', '8. Act']
    primera_fila = df_decprod.iloc[0]
    primera_fila = pd.DataFrame(primera_fila).transpose()

    # Escribir el DataFrame en el archivo Excel
    primera_fila.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False, header=is_empty)

    # Guardar el archivo Excel
    writer.save()
    print('procesing...', imagen)