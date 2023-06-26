from PIL import Image, ImageEnhance
import pytesseract
from PIL import Image
import pandas as pd
import datetime
import os
from openpyxl import load_workbook
import re
import numpy as np
import string
import dateutil.parser as dparser
import unidecode



pytesseract.pytesseract.tesseract_cmd = 'C:\\Program Files\\Tesseract-OCR\\tesseract.exe'

def write(ruta, folder, numb, contador):
    # especificar la ruta de la imagen
    imagen_path = os.path.join(ruta, folder, numb, contador)
    img = Image.open(imagen_path)
    contador_numerico = ''.join(filter(str.isdigit, contador))

    # convertir la imagen a texto utilizando PyTesseract, se deja en global para cargar la variable texto en la funcion cargarExcel
    texto = pytesseract.image_to_string(img, lang='spa') 
    if not texto:
        texto = 'x'
    texto = (texto.replace("~", "").replace("-","").replace("|", "").replace("?","").replace("(", "").replace(")", "").replace(";-","").replace('"', "").replace("]", "").replace("[", "").replace("5{", "").replace(";}", "").replace(";-", "").replace("-", "").replace(" /","").replace("aaa","")
        .replace("  ","").replace(" . “A/S \ ","").replace("fo","").replace("/","").replace("    ","").replace("&","").replace("N\ ","").replace("‘","").replace("\\",""))
    texto = unidecode.unidecode(texto)
    texto = re.sub(r"[^a-zA-ZñÑ0-9\s,.]++", " ", str(texto))  
    # Obtener la fecha actual y formatearla como YYYY-MM-DD
    fecha_actual = datetime.datetime.now().strftime("%Y_%m_%d")

    # Combinar la ruta, folder y numb para obtener la ruta completa del archivo Excel
    if numb == 'full':
        seachtype= os.path.join('input',folder,'TIPO BUSQUEDA.txt')
        try:
            with open(seachtype, 'r') as f:
                first_word = f.readline().split()[0]
        except FileNotFoundError:
            # Si el archivo no se encuentra, asigna una cadena vacía a first_word
            first_word = ""
        archivo_txt_path = os.path.join(ruta, folder,f"{contador_numerico}.txt")
        if "avanzado" in first_word:
            print("Busqueda avanzada en: ",folder)
            # Definir tamaño de corte (4cm de altura)
            corte_alto = 4 * 96  # 96 píxeles por cm
            if folder == '3. CEDULA':
                corte_alto= 8*96
            texto_total = ''

            for y in range(0, img.height, corte_alto):
                # Calcular límites de recorte
                y_inicio = y
                y_fin = min(y + corte_alto, img.height)

                # Recortar sección de 5cm de alto y ancho completo
                seccion = img.crop((0, y_inicio, img.width, y_fin))

                # Leer el texto de la sección utilizando pytesseract
                texto = pytesseract.image_to_string(seccion, lang='spa')

                # Procesar y manipular el texto
                texto_limpio = re.sub(r'[^a-zA-Z0-9,. ]+', ' ', texto)
                palabras = re.findall(r'[a-zA-Z]+|[0-9]+(?:[,.][0-9]+)?', texto_limpio)
                texto_total += ' '.join(palabras) + '\n'

            # Guardar el texto total en un archivo de texto
            with open(archivo_txt_path, 'w') as archivo_txt:
                archivo_txt.write(texto_total)

        else:
            with open(archivo_txt_path, 'w') as archivo_txt:
                archivo_txt.write(texto)
    else:
        # Combinar la ruta, folder y numb para obtener la ruta completa del archivo Excel
        
        archivo_excel_path = os.path.join(ruta, folder,f"{numb}_{fecha_actual}.xlsx")

        # Inicializar la variable libro_trabajo
        libro_trabajo = None
        contador = contador.split(".")[0]
        # Verificar si el archivo Excel ya existe
        if os.path.exists(archivo_excel_path):
            # Si el archivo Excel ya existe, cargar el libro de trabajo existente
            libro_trabajo = load_workbook(archivo_excel_path)

            # Seleccionar la hoja activa del libro de trabajo
            hoja_activa = libro_trabajo.active

            # Buscar la fila correspondiente al contador
            for fila in hoja_activa.iter_rows(min_row=2, max_col=1, values_only=True):
                if fila[0] == int(contador):
                    # Si se encuentra el contador, agregar el texto a la celda de la columna correspondiente
                    columna = chr(ord('A') + len(hoja_activa[1]))
                    print(fila[0],columna,len(hoja_activa[1]))
                    hoja_activa[f"{columna}{int(fila[0])}"] = texto
                    break
            else:
                # Si no se encuentra el contador, agregar una nueva fila con el contador y el texto
                nueva_fila = [int(contador), texto]
                hoja_activa.append(nueva_fila)

        else:
            # Si el archivo Excel no existe, crear un nuevo dataframe con el contador y el texto
            df = pd.DataFrame({"Contador": [int(contador)], numb: [texto]})

            # Guardar el dataframe en un nuevo archivo Excel
            df.to_excel(archivo_excel_path, index=False)

            # Cargar el libro de trabajo nuevo
            libro_trabajo = load_workbook(archivo_excel_path)

        # Guardar los cambios en el archivo Excel y cerrar el libro de trabajo
        libro_trabajo.save(archivo_excel_path)
        libro_trabajo.close()

        # Imprimir mensaje de confirmación
        print(f"Los datos han sido agregados al archivo Excel {numb}_{fecha_actual}.xlsx en la fila correspondiente al contador {contador}")
        

        # Imprimir mensaje de confirmación

def merge_excel_files(ruta, folder, folder1):
    
    if folder1 == "":
        excel_files_out=os.path.join(ruta)
    else:
        excel_files_out=os.path.join(ruta,folder1)
    
    fecha_actual = datetime.datetime.now().strftime("%d_%m_%Y")
    # Obtener la lista de archivos Excel en la carpeta especificada
    try: 
        excel_files = [f for f in os.listdir(os.path.join(ruta, folder)) if f.endswith('.xlsx')]
    except FileNotFoundError:
        # si el archivo no se encuentra, asignar el valor predeterminado a lineas
        excel_files= []


    # Verificar que hay al menos dos archivos Excel para combinar
    if len(excel_files) < 2:
        print("No hay suficientes archivos Excel para combinar.")
        return

    # Inicializar un diccionario para almacenar los dataframes de cada archivo Excel
    dfs = {}

    # Iterar sobre los archivos Excel y cargar cada uno como un dataframe en el diccionario
    # Iterar sobre los archivos Excel y cargar cada uno como un dataframe en el diccionario
    for file in excel_files:
        # Obtener el número de la columna a partir del nombre del archivo Excel
        numb = file.split('_')[0]

        # Cargar el archivo Excel como un dataframe y almacenarlo en el diccionario
        df = pd.read_excel(os.path.join(ruta, folder, file), index_col=0)
        df.reset_index(inplace=True)
        dfs[numb] = df


    # Concatenar los dataframes en uno solo con la función pd.concat()
    # El parámetro axis=1 indica que los dataframes se deben concatenar por columnas
    combined_df = pd.concat(dfs.values(), axis=1)

    # Crear un nuevo archivo Excel con el dataframe combinado
    combined_df.to_excel(os.path.join(excel_files_out, f"{folder}_{fecha_actual}.xlsx"), index=True)

    # Imprimir mensaje de confirmación
    print(f"archivos {combined_df}.xlsx")

def WrExcl(xlsx_name1,nombre,apell_str,CC,act_str,cant_str,uni_str,mun_str,dia_str,im):
    df2 = pd.DataFrame({'NOMBRE_1':[nombre],'APELLIDO_1':[apell_str],'CEDULA_1':[CC],'CODIGO_ACT_ECONOMICA':[act_str],'CANTIDAD_DE_MINERAL_VENDIDA':[cant_str],'UNIDAD_DE_MEDIDA':[uni_str],'MUNICIPIO_ORIGEN':[mun_str],'DIA':[dia_str],'IMG':[im]})
    writer = pd.ExcelWriter(xlsx_name1, engine='openpyxl', mode='a', if_sheet_exists='overlay')
    df2.to_excel(writer,index=False ,sheet_name='Sheet1', startrow=writer.sheets['Sheet1'].max_row, header=None)
    writer.save()
    writer.close()

def lector(excel_or,rutatxt):
    inputcol= os.path.join('input',rutatxt,'COLUMNAS.txt')
    lineas_default = ['1. Nombre', '2. Apellido', '3. Cc']
    rutadir= os.path.join('output',rutatxt)
    rutaerror= os.path.join('output','errores')

    # Carga el archivo de Excel en un DataFrame
    df = pd.read_excel('output\\salida\\'+excel_or)
    df_txt = pd.DataFrame()
    columnas_or = df.columns.to_list()

    # Carga el archivo de Excel en un DataFrame
    df = pd.read_excel('output\\salida\\'+excel_or)
    df_txt = pd.DataFrame()
    columnas_or = df.columns.to_list()

    try:
        with open(inputcol, 'r') as f:
            # leer todo el contenido del archivo
            contenido = f.read()
            # dividir el contenido en líneas utilizando el carácter de nueva línea como separador
            lineas = contenido.split('\n')
            if rutatxt == '3. CEDULA':
                lineas = lineas_default
    except FileNotFoundError:
        # si el archivo no se encuentra, asignar el valor predeterminado a lineas
        lineas = lineas_default
    except:
        # si se produce un error diferente, imprimir un mensaje de error y asignar el valor predeterminado a lineas
        print("Se produjo un error al leer el archivo.")
        lineas = lineas_default

    #listar txt.
    List_txt=[f for f in os.listdir(rutadir) if f.endswith('.txt')] 
    result=[]
    df = pd.read_excel('output\\salida\\'+excel_or)
    # Define el patrón de fecha esperado
    pattern = r'(\d{2})\n+(\d{2})\n+(\d{4})'


    # Convierte la columna '7. Fecha' al formato esperado
    df['7. Fecha'] = df['7. Fecha'].apply(lambda x: '/'.join(re.findall(pattern, x)[0]) if isinstance(x, str) and len(re.findall(pattern, x))>0 else x)

    # Convierte la columna '7. Fecha' al formato de fecha de Pandas y, en caso de no ser posible, deja los valores como estaban
    df['7. Fecha'] = pd.to_datetime(df['7. Fecha'], format='%d/%m/%Y', errors='coerce').fillna(df['7. Fecha'])


    df['3. Cc'] = df['3. Cc'].replace(regex=r'\D', value='')
    df['4. Cantidad'] = df['4. Cantidad'].apply(lambda x: re.sub(r'[^\d,.\s]+', '', x))
    df['4. Cantidad'] = df['4. Cantidad'].apply(lambda x: "x" if x.strip() == "" else str(float(x)) if x.isdigit() else x.replace(" ", "x"))
    df['4. Cantidad'] = df['4. Cantidad'].apply(lambda x: float(x) if (isinstance(x, str) and x.isdigit() and 10 <= float(x) <= 40) else x)

    # Convierte la columna del DataFrame a tipo 'str'
    df = df.astype(str)

    # Guarda el DataFrame modificado en un archivo Excel
    df.to_excel('output\\salida\\' + excel_or, index=False)

    # Imprime el DataFrame modificado
    df_out = pd.DataFrame()

    # Lista de caracteres de puntuación
    punctuation_list = list(string.punctuation)
    punctuation_list.append(' ')

    for titulo in lineas:
        try:
            result = []
            for t in df[titulo].to_list():
                comp = 'x'
                if pd.isna(t):
                    continue
                for txt in List_txt:
                    with open(rutadir + '\\' + txt, 'r') as file:
                        tx = t.strip().lower()
                        filex = file.read().lower()

                        # Eliminar la puntuación del texto
                        tx_no_punc = tx.translate(str.maketrans('', '', ''.join(punctuation_list)))
                        filex_no_punc = filex.translate(str.maketrans('', '', ''.join(punctuation_list)))

                        # Verificar si al menos 7 caracteres del texto se encuentran en el archivo con o sin puntuación
                        if any(filex[i:i+7] in tx or filex_no_punc[i:i+7] in tx_no_punc for i in range(len(filex)-6)):
                            comp = t
                    file.close()
                result.append(comp)
            df_out[rutatxt + '_' + titulo] = result
        except:
            print("Salto")

    fecha = [] # lista para almacenar todas las fechas encontradas

    if rutatxt+'_7. Fecha' in df_out.columns:
        # Diccionario para almacenar los valores encontrados en cada archivo
        archivos = {}
        patron_archivo = re.compile(r'\d+\.txt')

        for filename in sorted(filter(patron_archivo.match, os.listdir(rutadir)), key=lambda x: int(x.split('.')[0])):
            if filename.endswith(".txt"):
                # Inicializa una lista vacía para almacenar los valores de línea
                archivos[filename] = []
                # Abre el archivo de texto en modo lectura
                with open(os.path.join(rutadir, filename), "r") as archivo:
                    # Ciclo que recorre todas las líneas del archivo
                    for linea in archivo:
                        if re.search(r'\d{4}', linea):
                            # Verifica si la línea contiene un año (de 1000 a 2999)
                            if any(str(year) in linea for year in range(2019, 2050)):
                                # agrega la línea al diccionario correspondiente al archivo actual
                                archivos[filename].append(linea.strip())
                        

        # Ciclo para concatenar los valores encontrados en cada archivo
        for filename, valores in archivos.items():
            fecha.append(','.join(valores))

        # Crea un dataframe a partir de la lista de fechas
        

        df_out[rutatxt+'_7. Fecha']= fecha 
                        
    else:
        print("La columna '7. Fecha' esta limitada.")
        #Diccionario para almacenar los valores encontrados en cada archivo
        
        archivos = {}
        patron_archivo = re.compile(r'\d+\.txt')
        try:
            with open('input\\'+rutatxt+'\\FECHA.txt', 'r') as f:
                # leer todo el contenido del archivo
                contenido = f.read()
        except FileNotFoundError:
                # Si el archivo no se encuentra, asigna una cadena vacía a first_word
                contenido = '8. Fecha inicio:ddmmaaaa\n9. Fecha fin: ddmmaaaa'
            # dividir el contenido en líneas utilizando el carácter de nueva línea como separador
        txtfecha = contenido.split('\n')
            # crear una lista de folios

        # Separar el primer elemento de la txtfecha
        partes = txtfecha[0].split(':')
        fecha_inicio = partes[1].strip()

        # Separar el segundo elemento de la txtfecha
        partes = txtfecha[1].split(':')
        fecha_fin = partes[1].strip()

        # Imprimir los resultados
        fecha_inicio=fecha_inicio.replace('dd', r'(\d{2})').replace('mm', r'(\w+)').replace('aaaa', r'(\d{4})')
        fecha_fin = fecha_fin.replace('dd', r'(?P<dia>\d{2})\s*').replace('mm', r'(?P<mes>\w+)\s*').replace('aaaa', r'(?P<anio>\d{4})\s*')


        for filename in sorted(filter(patron_archivo.match, os.listdir(rutadir)), key=lambda x: int(x.split('.')[0])):
            if filename.endswith(".txt"):
                #Inicializa una lista vacía para almacenar los valores de línea
                archivos[filename] = []
                #Abre el archivo de texto en modo lectura
                with open(os.path.join(rutadir, filename), "r") as archivo:
                    #Ciclo que recorre todas las líneas del archivo
                    for linea in archivo:
                        if re.search(r'\d{4}', linea):
                            # Eliminar puntos y comas
                            linea = linea.replace('.', '').replace(',', '')
                            # Verifica si la línea contiene un año (de 1000 a 2999)
                            if any(str(year) in linea for year in range(1910, 2050)):
                                # Agrega la línea al diccionario correspondiente al archivo actual
                                archivos[filename].append(linea.strip())



        #Ciclo para concatenar los valores encontrados en cada archivo
        for filename, valores in archivos.items():

            
            fecha.append(' '.join(valores))
        if rutatxt == '3. CEDULA':
            patternini = r'([1-9]|[12]\d|3[01])\s+(ene|feb|mar|abr|may|jun|jul|ago|sep|oct|nov|dic)\s+(19[1-9]\d|20[0-4]\d|2050)'
            patternfin= r'([1-9]|[12]\d|3[01])\s+(ene|feb|mar|abr|may|jun|jul|ago|sep|oct|nov|dic)\s+(19[1-9]\d|20[0-4]\d|2050)'
        else:
            patternini = (fecha_inicio)
            patternfin= (fecha_fin)
            
        ini=['x']*len(fecha)
        fin=['x']*len(fecha)  # Crear lista inicialmente llena de 'x'
        for i, f in enumerate(fecha):  # Utilizar enumerate para tener índice i en el loop
            matchin = re.search(patternini, f, re.IGNORECASE)
        
            if matchin:
                diain, mesin, anioin = matchin.groups()
                fecha_strin = f"{diain}/{mesin}/{anioin}"
                ini[i] = fecha_strin
                f = f.replace(matchin.group(0), '')  # reemplazar la subcadena que coincide con el patrón por una cadena vacía
                
                matchfin= re.search(patternfin, f, re.IGNORECASE)      
                if matchfin:
                    dia, mes, anio = matchfin.groups()
                    fecha_str = f"{dia}/{mes}/{anio}"
                    fin[i] = fecha_str  # Reemplazar elemento en índice i de la lista 'fin'

            else:
                fin[i]=ini[i]


        df_out[rutatxt+'_8. Fecha inicio']= ini
        df_out[rutatxt+'_9. Fecha fin']= fin

    df_out.to_excel('output\\salida\\'+rutatxt+'.xlsx', index=False)



def mergemaster(output,salida):
 
    # Directorio donde se encuentran los archivos Excel
    directorio = os.path.join(output,salida)
    fecha_actual = datetime.datetime.now().strftime("%d_%m_%Y")

    # Lista para almacenar los nombres de los archivos Excel
    archivos_excel = []

    # Recorre todos los archivos del directorio y selecciona solo los que tienen extensión .xlsx
    for archivo in os.listdir(directorio):
        if archivo.endswith(".xlsx"):
            archivos_excel.append(archivo)

    # Concatena las columnas de todos los archivos Excel en un solo DataFrame
    df_concatenado = pd.DataFrame()
    for archivo in archivos_excel:
        df_excel = pd.read_excel(os.path.join(directorio, archivo))
        df_concatenado = pd.concat([df_concatenado, df_excel], axis=1)

    # Guarda el DataFrame concatenado en un nuevo archivo Excel
    df_concatenado.to_excel(output+"/"+salida+fecha_actual+".xlsx", index=False)

def definir(enrutador,enrutador2,colum):

    fecha_actual = datetime.datetime.now().strftime("%d_%m_%Y")
    xlsx_name1='output\\salida\\'+enrutador+'_'+fecha_actual+'.xlsx'# ruta del archivo output para el primer pdf
    xlsx_name2='output\\salida\\'+enrutador2+'_'+fecha_actual+'.xlsx'# ruta del archivo output para el primer pdf

    # Leer los archivos Excel
    try:
        df1 = pd.read_excel(xlsx_name1)
        df2 = pd.read_excel(xlsx_name2)
        
    except FileNotFoundError:
        # Si el archivo no se encuentra
        print("Aun no se ha genenrado el excel")



    df2['3. Cc'] = df1['3. Cc'].astype(str)

    df2['5. Act_ppa'] = df2['5. Act_ppa'].astype(str)

    df2['3. Cc'] = df2['3. Cc'].apply(lambda x: re.sub(r'[^\d,.\s]+', '', x))
    df2['3. Cc'] = df2['3. Cc'].apply(lambda x: "x" if x.strip() == "" else str(int(x)) if x.isdigit() else x.replace(" ", "x"))
    df2['5. Act_ppa'] = df2['5. Act_ppa'].apply(lambda x: re.sub(r'[^\d,.\s]+', '', x))
    df2['5. Act_ppa'] = df2['5. Act_ppa'].apply(lambda x: "x" if x.strip() == "" else str(float(x)) if x.isdigit() else x.replace(" ", "x"))
    if colum=='3. Cc:' :
        df2['3. Cc']=df1[colum]
    df1[colum] = df1[colum].fillna('x')
    df2[colum] = df2[colum].fillna('x')

    pattern = r'(\d{4})(\d{2})(\d{2})'
    # Reemplaza los espacios en blanco en la columna '10. Fecha rut'
    df2['10. Fecha rut'] = df2['10. Fecha rut'].replace(' ', '', regex=True)

    try:
        # Convierte la columna '10. Fecha rut' al formato esperado
        df2['10. Fecha rut'] = df2['10. Fecha rut'].apply(lambda x: '/'.join(re.findall(pattern, x)[0]) if isinstance(x, str) and len(re.findall(pattern, x))>0 else x)

        # Convierte la columna '10. Fecha rut' al formato de fecha de Pandas y, en caso de no ser posible, deja los valores como estaban
        df2['10. Fecha rut'] = pd.to_datetime(df2['10. Fecha rut'], format='%d/%m/%Y', errors='coerce').fillna(df2['10. Fecha rut'])

        # Obtiene solo la fecha de la columna '10. Fecha rut' y la convierte a una cadena de texto
        df2['10. Fecha rut'] = df2['10. Fecha rut'].dt.date.astype(str)
    except AttributeError:
        # Maneja la excepción y continua con el código
        print("Se produjo una excepción AttributeError. El programa continuará sin convertir la columna '10. Fecha rut'")


    # Convertir columnas en listas para comparar elemento a elemento
    lista1 = df1['1. Nombre'].tolist()
    lista2 = df2['1. Nombre'].tolist()

    # Recorremos las dos listas al mismo tiempo
    for i in range(len(lista1)):
        # Buscamos el índice del primer carácter que coincide
        indice = lista1[i].find(lista2[i][0])
        # Si se encuentra el primer carácter, verificamos si hay al menos 4 caracteres consecutivos iguales
        if indice != -1 and lista1[i][indice:indice+4] == lista2[i][:4]:
            # Reemplazamos la fila correspondiente en la lista 1 con la fila correspondiente en la lista 2
            lista2[i] = lista1[i] + lista2[i][indice+4:]
    # Convertir listas de nuevo a dataframes
    df2['1. Nombre'] = lista2


    # Escribir los archivos actualizados
    df1.to_excel(xlsx_name1, index=False)
    df2.to_excel(xlsx_name2, index=False)

def unir_contador_columnas(excel_salida):
    # Cargar el archivo de Excel en un dataframe
    df = pd.read_excel(excel_salida)

    # Eliminar todas las columnas que se llaman 'Unnamed'
    df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
    df = df.loc[:, ~df.columns.str.contains('^Contador')]

    # Guardar los datos modificados en un nuevo archivo de Excel
    writer = pd.ExcelWriter(excel_salida)
    df.to_excel(writer, index=False)
    writer.save()

def unificarCEDULA(enrutador, enrutador2):
    fecha_actual = datetime.datetime.now().strftime("%d_%m_%Y")
    xlsx_name1='output\\salida\\'+enrutador+'_'+fecha_actual+'.xlsx'# ruta del archivo output para el primer pdf
    xlsx_name2='output\\salida\\'+enrutador2+'.xlsx'# ruta del archivo output para el primer pdf
    # Leer los archivos Excel
    df1 = pd.read_excel(xlsx_name1)
    df2 = pd.read_excel(xlsx_name2)

    df1['3. Cc'] = df1['3. Cc'].astype(str)
    df2['3. CEDULA_3. Cc'] = df2['3. CEDULA_3. Cc'].astype(str)

    # Seleccionar las filas en df2 que necesitan ser actualizadas
    maskname = df2['3. CEDULA_1. Nombre'].isin(df1['1. Nombre'])
    maskcc = df2['3. CEDULA_3. Cc'].isin(df1['3. Cc'])

    # Actualizar los valores en df2 con los valores correspondientes en df1
    df2.loc[maskcc, '3. CEDULA_1. Nombre'] = df1['1. Nombre']
    df2.loc[maskcc, '3. CEDULA_2. Apellido'] = df1['2. Apellido']
    df2.loc[maskname, '3. CEDULA_2. Apellido'] = df1['2. Apellido']
    df2.loc[maskname, '3. CEDULA_3. Cc'] = df1['3. Cc']

    # Escribir los datos actualizados en el archivo Excel
    df2.to_excel(xlsx_name2, index=False)